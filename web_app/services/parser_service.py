"""
文件解析服务 — PPTX / PDF / XLSX
"""
import re
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path

import pdfplumber
import openpyxl


def parse_pptx(filepath: str) -> dict:
    """解析 PPTX，返回结构化的文本内容"""
    slides_data = []
    with zipfile.ZipFile(filepath, 'r') as z:
        slide_names = sorted([
            n for n in z.namelist()
            if n.startswith('ppt/slides/slide') and n.endswith('.xml')
        ], key=lambda x: int(re.search(r'slide(\d+)', x).group(1)))
        for sn in slide_names:
            slide_num = int(re.search(r'slide(\d+)', sn).group(1))
            with z.open(sn) as f:
                tree = ET.parse(f)
                root = tree.getroot()
                texts = []
                for t in root.iter('{http://schemas.openxmlformats.org/drawingml/2006/main}t'):
                    if t.text:
                        texts.append(t.text)
            if texts:
                slides_data.append({
                    "slide": slide_num,
                    "content": '\n'.join(texts)
                })
    full_text = '\n\n'.join([s['content'] for s in slides_data])
    return {
        "slides": slides_data,
        "full_text": full_text,
        "slide_count": len(slides_data)
    }


def parse_pdf(filepath: str) -> dict:
    """解析 PDF，返回结构化的文本内容"""
    pages_data = []
    with pdfplumber.open(filepath) as pdf:
        for i, page in enumerate(pdf.pages):
            text = page.extract_text()
            if text:
                pages_data.append({"page": i + 1, "content": text})
    full_text = '\n\n'.join([p['content'] for p in pages_data])
    return {"pages": pages_data, "full_text": full_text}


def parse_xlsx(filepath: str) -> dict:
    """解析 XLSX，返回结构化的文本内容"""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheets_data = {}
    for sn in wb.sheetnames:
        ws = wb[sn]
        rows = []
        for row in ws.iter_rows(values_only=True):
            vals = [str(c) if c is not None else '' for c in row]
            if any(v.strip() for v in vals):
                rows.append(' | '.join(vals))
        sheets_data[sn] = '\n'.join(rows)
    full_text = '\n\n'.join(sheets_data.values())
    return {"sheets": sheets_data, "full_text": full_text}


def _extract_level(filename: str, full_text: str) -> str:
    """从文件名和内容中提取申报级别，兼容大小写/全角横线/空格等写法
    返回规范格式如 S3-2、T1；提取不到返回空串"""
    dash = r'[-‐‑–—－﹣]'
    # 优先：文件名中的完整级别（如 申请S4-3 / s3—2）
    m = re.search(r'([SPTspt])\s*(\d)\s*' + dash + r'\s*(\d)', filename)
    if m:
        return f"{m.group(1).upper()}{m.group(2)}-{m.group(3)}"
    # 其次：内容中"申请/申报/职级/级别"附近的完整级别
    m = re.search(r'(?:申请|申报|职级|级别)[^\n]{0,30}?([SPTspt])\s*(\d)\s*' + dash + r'\s*(\d)', full_text[:5000])
    if m:
        return f"{m.group(1).upper()}{m.group(2)}-{m.group(3)}"
    # 内容中任意位置的完整级别（如 拟认证岗位与职级：运营专员+S3-2）
    m = re.search(r'([SPTspt])(\d)' + dash + r'(\d)', full_text[:5000])
    if m:
        return f"{m.group(1).upper()}{m.group(2)}-{m.group(3)}"
    # 最后：文件名中的单级别写法（如 申请S1，助理级无档位）
    m = re.search(r'(?:申请|申报)\s*([SPTspt])(\d)(?!\d)', filename)
    if m:
        return f"{m.group(1).upper()}{m.group(2)}"
    return ''


def _clean_department(dept: str) -> str:
    """清洗部门字段：去掉括号注释（如"(已瘦身)"），只保留第一段业务单元名
    示例："(已瘦身)傲点—产品推广部" → "傲点" """
    cleaned = re.sub(r'[（(][^）)]*[）)]', '', dept).strip()
    cleaned = re.split(r'[-‐‑–—－﹣]', cleaned)[0].strip()
    return cleaned if cleaned else dept


def _clean_position(pos: str) -> str:
    """清洗申报岗位字段：去掉"（申请Sx-x"级别后缀（含被横线截断的残缺括号）
    示例："高级GTM（申请S4" → "高级GTM"；"申请落位S3" → ""（无有效岗位名）"""
    cleaned = re.split(r'[（(]\s*申[请报]', pos)[0].strip()
    cleaned = re.sub(r'[（(][^）)]*$', '', cleaned).strip()          # 未闭合括号残段
    cleaned = re.sub(r'[-‐‑–—－﹣]?\s*[SPTspt]\d(-\d)?\s*[）)]?$', '', cleaned).strip()  # 粘连的级别代号
    if not cleaned or re.match(r'^(申请|申报|落位)', cleaned):
        return ''
    return cleaned


def extract_employee_info(full_text: str, filename: str) -> dict:
    """从PPT内容和文件名中提取员工信息"""
    info = {
        "员工姓名": "未知",
        "所在部门": "未知",
        "当前岗位": "未知",
        "申报岗位": "未知",
        "申报级别": "未知",
        "汇报日期": "未知",
    }

    # 从文件名提取: 部门-姓名-申报岗位（申请级别）
    name_part = filename.rsplit('.', 1)[0]
    parts = [p.strip() for p in name_part.split('-')]
    if len(parts) >= 2:
        info["所在部门"] = parts[0] if parts[0] else "未知"
        info["员工姓名"] = parts[1] if len(parts) > 1 else "未知"
    if len(parts) >= 3:
        info["申报岗位"] = parts[2] if len(parts) > 2 else "未知"

    level_match = _extract_level(filename, full_text)
    if level_match:
        info["申报级别"] = level_match

    # 从内容提取
    for key, pattern in [
        ("员工姓名", r'汇报人[：:]?\s*(.+)'),
        ("员工姓名", r'姓名[：:]\s*(.+)'),
        ("所在部门", r'所在部门[：:]\s*(.+)'),
        ("申报岗位", r'拟认证岗位[与和]职级[：:]\s*(.+)'),
        ("申报岗位", r'当前岗位[与和]职级[：:]\s*(.+)'),
        ("汇报日期", r'日\s*期[：:]\s*(\d{4}[./]\d{1,2})'),
        ("汇报日期", r'(\d{4}[./]\d{1,2})'),
    ]:
        m = re.search(pattern, full_text)
        if m and info[key] in ("未知", ""):
            info[key] = m.group(1).strip()

    m = re.search(r'拟认证岗位[与和]职级[：:]\s*【?(.+?】?)?[+＋]\s*([SPT]\d-\d)', full_text)
    if m:
        info["申报岗位"] = m.group(1).strip('【】 ') if m.group(1) else info["申报岗位"]
        info["申报级别"] = m.group(2)

    # 部门清洗：去括号注释、只留业务单元名（如"(已瘦身)傲点—产品推广部" → "傲点"）
    if info["所在部门"] not in ("未知", ""):
        info["所在部门"] = _clean_department(info["所在部门"])

    # 岗位清洗：去掉"（申请Sx-x"级别后缀；清洗后无有效岗位名的置为未知
    for key in ("申报岗位", "当前岗位"):
        if info[key] not in ("未知", ""):
            info[key] = _clean_position(info[key]) or "未知"

    return info


def extract_standard_summary(filepath) -> dict:
    """从标准XLSX中提取摘要信息：职责/场景清单、覆盖部门"""
    wb = openpyxl.load_workbook(str(filepath), data_only=True)
    duty_names = []
    departments = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(values_only=True):
            vals = [str(c).strip() if c is not None else '' for c in row]
            if not any(vals):
                continue
            first = vals[0]
            # 职责1 / 场景1 行 → 第二列为职责名称
            if re.match(r'^(职责|场景)\s*\d+$', first):
                name = next((v for v in vals[1:] if v), '')
                name = name.split('\n')[0].strip()[:50]
                if name and name not in duty_names:
                    duty_names.append(name)
            # 覆盖部门行
            elif first.startswith('覆盖部门'):
                dept_text = next((v for v in vals[1:] if v), '')
                for d in re.split(r'[、，,;；/\s]+', dept_text):
                    d = d.strip()
                    if d and d not in departments:
                        departments.append(d)
        # 第一个解析出职责的sheet即为主sheet，避免多sheet重复统计
        if duty_names:
            break

    return {"duty_names": duty_names, "departments": departments}


def find_standard_file(standards_dir, filename: str):
    """在岗位标准目录下查找标准文件（支持P序列/S序列/T序列等子文件夹）"""
    direct = Path(standards_dir) / filename
    if direct.exists():
        return direct
    for f in Path(standards_dir).rglob('*'):
        if f.is_file() and f.name == filename:
            return f
    return None


def match_standard(filename: str, full_text: str, registry_file: Path) -> list:
    """从注册表中匹配岗位标准"""
    import json
    registry = json.loads(registry_file.read_text(encoding='utf-8'))
    combined = (filename + ' ' + full_text[:2000]).lower()

    # 从文件名解析申报岗位（格式：部门-姓名-申报岗位（申请级别）），用于反向包含匹配
    applied_pos = ''
    parts = [p.strip() for p in filename.rsplit('.', 1)[0].split('-')]
    if len(parts) >= 3:
        applied_pos = re.sub(r'[（(].*', '', parts[2]).strip()
        applied_pos = re.sub(r'^(助理|初级|中级|高级|资深|专家)', '', applied_pos).lower()

    # 从文件名提取申报级别的序列字母（如 S4-3 → S），用于同分时优先本序列
    seq_letter = ''
    seq_m = re.search(r'([SPT])\d-\d', filename.upper())
    if seq_m:
        seq_letter = seq_m.group(1)

    scores = []
    for std in registry["岗位标准清单"]:
        name = std["岗位名称"].lower()
        score = 0
        for kw in std.get("关键词", []):
            if kw.lower() in combined:
                score += 1
        if name in combined:
            # 完整岗位名命中
            score += 3
        else:
            # 基础名（去括号后缀）命中，如 "运营专员（电商）" → "运营专员"
            base = re.split(r'[（(]', name)[0].strip()
            if base and base in combined:
                score += 2
                # 括号内标签同时命中则加分，用于区分同基础名的标准
                for tag in re.findall(r'[（(]([^）)]+)[）)]', name):
                    if tag.lower() in combined:
                        score += 1
            # 反向包含：文件名中的申报岗位是标准名的一部分，如 "运营" ⊆ "运营专员（电商）"
            if applied_pos and len(applied_pos) >= 2 and applied_pos in name:
                score += 2
            elif applied_pos and len(applied_pos) >= 3:
                # 兜底：申报岗位与标准名的二字词重叠（处理简称，如 "社媒运营" → "社媒&社群运营专员"）
                bigrams = [applied_pos[i:i+2] for i in range(len(applied_pos) - 1)]
                hits = sum(1 for bg in bigrams if bg in name)
                if hits / len(bigrams) >= 0.5:
                    score += 2
        # 同分时优先申报级别所属序列
        if score > 0 and seq_letter and std.get("序列", "").upper().startswith(seq_letter):
            score += 0.5
        if score > 0:
            scores.append((score, std))

    scores.sort(key=lambda x: -x[0])
    return [s[1] for s in scores]
